import glob
import re
from pathlib import Path
from langchain_community.document_loaders import UnstructuredWordDocumentLoader, PyPDFLoader
import pandas as pd
from openpyxl import load_workbook

"""
                        Load Excel File
-----------------------------------------------------------------------------------------------
"""
excel_path = "./data/appointments/appointments_technova.xlsx"
sheet_name = "Sheet1"
# load workbook
book = load_workbook(excel_path)
"""
                        Load Documents
-----------------------------------------------------------------------------------------------
"""
def file_reader(path: str, extensions: tuple):
    # give data path
    data_path = Path(path)

    # store all files (list of list)
    files = {}
    # match multiple patterns (e.x: txt, pdf) from subdirectories
    for ext in extensions:
        rglobed_files = (data_path.rglob(f"*.{ext.lower()}")) 
        files[ext] = [file.as_posix() for file in rglobed_files] # as_posix: convert path to standard Unix-style string

    # store all documents (unloaded)
    documents = []
    # initialize a default document loader
    document_loader = None
    for ext, file in files.items():
        if ext == "docx":
            document_loader = [UnstructuredWordDocumentLoader(ele) for ele in file]
        elif ext == "pdf":
            document_loader = [PyPDFLoader(ele) for ele in file]

        documents.extend(loader.load() for loader in document_loader)

    return documents

"""
                        Fetch Appointment Details
-----------------------------------------------------------------------------------------------
"""
def appointment_reader_saver(msg: str):
    # Use case-insensitive regex and allow flexible spacing/newlines
    patterns = {
        "name": r"(?i)name\s*[:\-]\s*([A-Za-z ]+)",
        "phone": r"(?i)phone(?: number)?\s*[:\-]\s*(\d{7,15})",
        "email": r"(?i)email(?: address)?\s*[:\-]\s*([\w\.-]+@[\w\.-]+\.\w+)",
        "date": r"(?i)date\s*[:\-]\s*([0-9]{4}-[0-9]{2}-[0-9]{2})"
    }

    extracted = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, msg, re.DOTALL)  # DOTALL makes it work across lines
        extracted[key] = [match.group(1).strip() if match else None]

    # dataframe to store appointment details
    df = pd.DataFrame(extracted)
    # create ExcelWriter object using openpyxl as engine
    # Set mode = 'a' for append mode and if_sheet_exists = 'overlay' to overwrite existing cells
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # Write the new DataFrame to the specified sheet
        # Set startrow to append data below existing content, or 0 to overwrite from the beginning
        # index=False prevents writing the DataFrame index to Excel
        df.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startrow=book[sheet_name].max_row)
